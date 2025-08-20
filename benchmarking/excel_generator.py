import pandas as pd
import glob
import os
from typing import Dict, List, Tuple, Optional
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings


def parse_summary_file(filepath: str) -> Dict[str, str]:
    """
    Parse a summary metrics file and extract metrics information.
    
    Args:
        filepath (str): Path to the summary metrics file
        
    Returns:
        Dict[str, str]: Dictionary containing parsed metrics
    """
    metrics = {}
    try:
        with open(filepath, 'r') as f:
            for line in f:
                line = line.strip()
                if ':' in line:
                    key, value = line.split(':', 1)
                    metrics[key.strip()] = value.strip()
    except Exception as e:
        warnings.warn(f"Error parsing {filepath}: {e}")
        return {}
    
    return metrics


def collect_summary_metrics(output_folder: str, pattern: str = "*summarymetrics.txt") -> pd.DataFrame:
    """
    Collect all summary metrics files from the output folder and parse them into a DataFrame.
    
    Args:
        output_folder (str): Directory containing summary files
        pattern (str): Glob pattern to match summary files
        
    Returns:
        pd.DataFrame: DataFrame with columns ['Model', 'Dataset', 'Score_Type', 'ROC_AUC', 'Average_Precision', 'F1_Score', 'Threshold']
    """
    search_pattern = os.path.join(output_folder, pattern)
    summary_files = glob.glob(search_pattern)
    
    if not summary_files:
        raise FileNotFoundError(f"No summary files found matching pattern: {search_pattern}")
    
    all_metrics = []
    
    for filepath in summary_files:
        metrics = parse_summary_file(filepath)
        if metrics:  # Only add if parsing succeeded
            all_metrics.append(metrics)
    
    if not all_metrics:
        raise ValueError("No valid summary files could be parsed")
    
    df = pd.DataFrame(all_metrics)
    
    # Convert numeric columns to float
    numeric_cols = ['ROC_AUC', 'Average_Precision', 'F1_Score', 'Threshold']
    for col in numeric_cols:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    
    return df


def create_pivot_table(df: pd.DataFrame) -> pd.DataFrame:
    """
    Create a pivot table with models as rows and dataset metrics as column groups.
    
    Args:
        df (pd.DataFrame): DataFrame from collect_summary_metrics
        
    Returns:
        pd.DataFrame: Pivot table formatted for Excel export
    """
    # Create a multi-level column structure
    pivot_data = []
    
    # Get unique datasets and models
    datasets = sorted(df['Dataset'].unique())
    models = sorted(df['Model'].unique())
    
    # Create column structure: Dataset -> Metric
    columns = []
    for dataset in datasets:
        for metric in ['ROC_AUC', 'Average_Precision', 'F1_Score']:
            columns.append((dataset, metric))
    
    # Create MultiIndex columns
    multi_columns = pd.MultiIndex.from_tuples(columns, names=['Dataset', 'Metric'])
    
    # Initialize the result DataFrame
    result_df = pd.DataFrame(index=models, columns=multi_columns)
    
    # Fill in the data
    for _, row in df.iterrows():
        model = row['Model']
        dataset = row['Dataset']
        
        for metric in ['ROC_AUC', 'Average_Precision', 'F1_Score']:
            if metric in row and pd.notna(row[metric]):
                result_df.loc[model, (dataset, metric)] = row[metric]
    
    return result_df


def rank_values_in_columns(df: pd.DataFrame) -> Dict[Tuple[str, str], List[Tuple[str, int]]]:
    """
    Rank values in each column and return ranking information for formatting.
    
    Args:
        df (pd.DataFrame): Pivot table with MultiIndex columns
        
    Returns:
        Dict: Dictionary mapping (dataset, metric) to list of (model, rank) tuples
    """
    rankings = {}
    
    for dataset, metric in df.columns:
        # Get the column data, excluding NaN values
        col_data = df[(dataset, metric)].dropna()
        
        if len(col_data) == 0:
            continue
            
        # Rank values (highest is rank 1)
        ranks = col_data.rank(method='min', ascending=False)
        
        # Create list of (model, rank) tuples
        model_ranks = [(model, int(rank)) for model, rank in ranks.items()]
        rankings[(dataset, metric)] = model_ranks
    
    return rankings


def format_excel_with_styling(df: pd.DataFrame, rankings: Dict[Tuple[str, str], List[Tuple[str, int]]], 
                              output_path: str) -> None:
    """
    Export DataFrame to Excel with custom formatting (bold for 1st place, italic for 2nd place).
    
    Args:
        df (pd.DataFrame): Pivot table to export
        rankings (Dict): Ranking information from rank_values_in_columns
        output_path (str): Path to save the Excel file
    """
    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Benchmarking Results"
    
    # Write headers
    # First row: Dataset names
    row = 1
    col = 2  # Start from column B (A is for model names)
    
    for dataset, metric in df.columns:
        ws.cell(row=row, column=col, value=dataset)
        col += 1
    
    # Second row: Metric names
    row = 2
    col = 2
    for dataset, metric in df.columns:
        ws.cell(row=row, column=col, value=metric)
        col += 1
    
    # Write model names and data
    row = 3
    for model_idx, model in enumerate(df.index):
        # Model name in column A
        ws.cell(row=row, column=1, value=model)
        
        # Data values
        col = 2
        for dataset, metric in df.columns:
            value = df.loc[model, (dataset, metric)]
            
            if pd.notna(value):
                # Round to 4 decimal places for display
                cell = ws.cell(row=row, column=col, value=round(float(value), 4))
                
                # Apply formatting based on ranking
                if (dataset, metric) in rankings:
                    model_ranks = dict(rankings[(dataset, metric)])
                    if model in model_ranks:
                        rank = model_ranks[model]
                        if rank == 1:
                            # Bold for first place
                            cell.font = Font(bold=True)
                        elif rank == 2:
                            # Italic for second place
                            cell.font = Font(italic=True)
            else:
                ws.cell(row=row, column=col, value="N/A")
            
            col += 1
        row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)  # Cap at 15 characters
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)


def generate_results_excel(output_folder: str, excel_filename: str = "benchmarking_results.xlsx", 
                          pattern: str = "*summarymetrics.txt") -> str:
    """
    Main function to generate Excel file with benchmarking results.
    
    This function:
    1. Searches for summary metrics files in the output folder
    2. Parses all metrics into a structured format
    3. Creates a pivot table with models as rows and datasets as column groups
    4. Ranks performance and applies formatting (bold for 1st, italic for 2nd)
    5. Exports to Excel with proper formatting
    
    Args:
        output_folder (str): Directory containing summary metrics files
        excel_filename (str): Name of the output Excel file
        pattern (str): Glob pattern to match summary files (default: "*summarymetrics.txt")
        
    Returns:
        str: Path to the generated Excel file
        
    Raises:
        FileNotFoundError: If no summary files are found
        ValueError: If no valid summary files could be parsed
    """
    # Collect and parse all summary files
    print(f"Collecting summary metrics from {output_folder}...")
    df = collect_summary_metrics(output_folder, pattern)
    print(f"Found {len(df)} summary files")
    
    # Create pivot table
    print("Creating pivot table...")
    pivot_df = create_pivot_table(df)
    print(f"Pivot table created with {len(pivot_df)} models and {len(pivot_df.columns)} metric columns")
    
    # Rank values for formatting
    print("Ranking values for formatting...")
    rankings = rank_values_in_columns(pivot_df)
    
    # Generate Excel file path
    excel_path = os.path.join(output_folder, excel_filename)
    
    # Export to Excel with formatting
    print(f"Exporting to Excel: {excel_path}")
    format_excel_with_styling(pivot_df, rankings, excel_path)
    
    print(f"✅ Excel file generated successfully: {excel_path}")
    return excel_path


def print_summary_stats(output_folder: str, pattern: str = "*summarymetrics.txt") -> None:
    """
    Print summary statistics about the collected metrics.
    
    Args:
        output_folder (str): Directory containing summary files
        pattern (str): Glob pattern to match summary files
    """
    try:
        df = collect_summary_metrics(output_folder, pattern)
        
        print("\n=== Summary Statistics ===")
        print(f"Total summary files found: {len(df)}")
        print(f"Unique models: {df['Model'].nunique()} ({', '.join(sorted(df['Model'].unique()))})")
        print(f"Unique datasets: {df['Dataset'].nunique()} ({', '.join(sorted(df['Dataset'].unique()))})")
        print(f"Unique score types: {df['Score_Type'].nunique()} ({', '.join(sorted(df['Score_Type'].unique()))})")
        
        # Show best performers by metric
        for metric in ['ROC_AUC', 'Average_Precision', 'F1_Score']:
            if metric in df.columns:
                best_idx = df[metric].idxmax()
                if pd.notna(best_idx):
                    best_row = df.loc[best_idx]
                    print(f"Best {metric}: {best_row['Model']} on {best_row['Dataset']} ({best_row[metric]:.4f})")
        
    except Exception as e:
        print(f"Error generating summary stats: {e}")


if __name__ == '__main__':
    # Example usage
    output_folder = "output_csvs"
    
    # Print summary statistics
    print_summary_stats(output_folder)
    
    # Generate Excel file
    try:
        excel_path = generate_results_excel(output_folder)
        print(f"\n✅ Benchmarking results exported to: {excel_path}")
    except Exception as e:
        print(f"❌ Error generating Excel file: {e}")